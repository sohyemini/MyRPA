import openpyxl
from openpyxl.styles import PatternFill

wb_ori = openpyxl.load_workbook(r".\files\address_org.xlsx")
ws_ori = wb_ori.active
mc_ori = ws_ori.max_column
mr_ori = ws_ori.max_row

flag1 = False
for i in range(2, mr_ori):
    for j in range(i+1, mr_ori+1):
        if ws_ori.cell(i, 1).value == ws_ori.cell(j, 1).value:
            print(f"원본파일 : {i}, {j}겹치는 키가 있습니다.")
            Flag1 = True
            break

if flag1 == False:
    print("원본 파일에는 겹치는 키가 없습니다.")

flag2 = False
wb_new = openpyxl.load_workbook(r".\files\address_new.xlsx")
ws_new = wb_new.active
mc_new = ws_new.max_column
mr_new = ws_new.max_row

for i in range(2, mr_new):
    for j in range(i+1, mr_new+1):
        if ws_new.cell(i, 1).value == ws_new.cell(j, 1).value:
            print(f"비교할 파일: {i}, {j}겹치는 키가 있습니다.")
            Flag2 = True
            break

if flag2 == False:
    print("비교할 파일에는 겹치는 키가 없습니다.")

if flag1 or flag2:
    print("중복 키가 있습니다.\n각 파일을 확인해주시기 바랍니다.")
    exit(0)

found = False
diff_color = PatternFill(start_color='ffff99', end_color='ffff99', fill_type='solid')
for i in range(2, mr_ori+1):        # 원본 키값 (행)
    for j in range(2, mr_new+1):    # 비교파일 키값 (행)
        if ws_ori.cell(i, 1).value == ws_new.cell(j, 1).value:  # 원본과 비교파일 키값이 같으면 (행)
            for k in range(2, mc_ori + 1): # 같은 키값들의 열을 비교함
                if ws_ori.cell(i,k).value != ws_new.cell(j, k).value: # (열)이 다르다면
                    if ws_ori.cell(i,k).value == None: ws_ori.cell(i,k).value = '' # None --> ''
                    if ws_new.cell(j,k).value == None: ws_new.cell(j,k).value = '' # None --> ''
                    ws_ori.cell(i,k).value += f"--> {ws_new.cell(j, k).value}"    # 바뀐값 표시
                    ws_ori.cell(i,k).fill = diff_color
            found = True
            break
    if found == False:
        print(f"원본 {i}행은 비교파일에서 삭제되었음")
    found = False

wb_ori.save(r".\files\address_diff.xlsx")
import openpyxl

wb_ori = openpyxl.load_workbook(r".\files\income_state_org.xlsx")
ws_ori = wb_ori['Income']
mc_ori = ws_ori.max_column
mr_ori = ws_ori.max_row

wb_new = openpyxl.load_workbook(r".\files\income_state_new.xlsx")
ws_new = wb_new['Income']
mc_new = ws_new.max_column
mr_new = ws_new.max_row

if mc_new != mc_ori:
    print("파일의 데이터 개수에 차이가 있습니다.")

for i in range(1, mr_ori+1, 1):
    print(i, end=' ')
    for j in range(1, mc_ori + 1, 1):
        print(ws_ori.cell(i, j).value, "|" ,ws_new.cell(i, j).value , end=' - ')
    print('')



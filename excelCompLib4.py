import openpyxl
from openpyxl.styles import PatternFill, Color

class StaticFormExcelCompare():
    def __init__(self):
        print("__init__")
        self.diff_color = "ffff99"
        self.big_color = "ffb399"
        self.small_color = "99ffb3"

    def setColor(self, diff, big, small):
        self.diff_color = diff
        self.big_color = big
        self.small_color = small

    def setFiles(self, ori, new, out):
        self.wb_ori = openpyxl.load_workbook(ori, data_only=True)
        self.wb_new = openpyxl.load_workbook(new, data_only=True)
        self.wb_out = out

    def run(self):
        ws_ori = self.wb_ori.active
        mc_ori = ws_ori.max_column
        mr_ori = ws_ori.max_row

        ws_new = self.wb_new.active
        mc_new = ws_new.max_column
        mr_new = ws_new.max_row

        if mc_new != mc_ori or mr_ori != mr_new:
            print("파일의 데이터 개수에 차이가 있습니다.")
            return False

        for i in range(1, mr_ori+1, 1):
            for j in range(1, mc_ori + 1, 1):
                if ws_ori.cell(i, j).value != ws_new.cell(i, j).value:
                    ori = ws_ori.cell(i, j).value
                    new = ws_new.cell(i, j).value
                    ws_ori.cell(i, j).value = f"{ws_ori.cell(i, j).value} --> {ws_new.cell(i, j).value}"

                    if isinstance(ori, str) or isinstance(new, str):
                        ws_ori.cell(i, j).fill = PatternFill(patternType='solid', fgColor=Color(self.diff_color))
                    else:
                        if ori > new:
                            ws_ori.cell(i, j).fill = PatternFill(patternType='solid', fgColor=Color(self.small_color))
                        else:
                            ws_ori.cell(i, j).fill = PatternFill(patternType='solid', fgColor=Color(self.big_color))

        try:
            self.wb_ori.save(self.wb_out)
            return True
        except:
            return False

a = StaticFormExcelCompare()
a.setColor('123456', '789ABC','DEF123')
a.setFiles(r".\files\income_state_org.xlsx", r".\files\income_state_new.xlsx", r".\files\income_state_comp2.xlsx")
if a.run():
    print("파일 비교 완료")
else:
    print("파일 비교 에러")


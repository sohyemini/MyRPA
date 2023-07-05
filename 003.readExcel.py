import openpyxl  # 패키지 import

wb = openpyxl.load_workbook(r".\files\codefilex.xlsx")  # 엑셀은 워크북(엑셀파일), 그리고 워크시트로 구분이 됩니다.

ws = wb['codefilex']

print(ws['A1'].value, ws.cell(1, 2).value)

print("프로그램이 종료되었습니다.")

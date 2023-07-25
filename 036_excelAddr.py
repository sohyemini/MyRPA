import openpyxl

wb = openpyxl.load_workbook(r".\files\codefilex.xlsx")
ws = wb['codefilex']

name = ''
postcode = ''
addr = ''

for i in range(2, 100): # 첫번째 줄은 제목이 나타남
    name = f"{ws.cell(i, 2).value} {ws.cell(i, 3).value} 귀하"
    postcode = ws.cell(i, 6).value
    addr = ws.cell(i, 7).value

    print(name, postcode, addr)

import openpyxl  # 패키지 import

wb = openpyxl.load_workbook(r".\files\codefilex.xlsx")
ws = wb['codefilex']

mc = ws.max_column
mr = ws.max_row

for i in range(1, mr+1, 1):
    print(i, end=' ')
    for j in range(1, mc + 1, 1):
        print(ws.cell(i, j).value, end=' ')
    print('')

print(f"이 엑셀 파일은 {mr}행, {mc}열로 이루어져 있습니다")
print("프로그램이 종료되었습니다.")

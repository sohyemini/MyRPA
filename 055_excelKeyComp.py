import openpyxl

wb_ori = openpyxl.load_workbook(r".\files\address_org.xlsx")
ws_ori = wb_ori.active
mc_ori = ws_ori.max_column
mr_ori = ws_ori.max_row

flag1 = False
for i in range(2, mr_ori):
    for j in range(i+1, mr_ori+1):
        if ws_ori.cell(i, 1).value == ws_ori.cell(j, 1).value:
            print(f"원본파일 : {i}, {j}겹치는 키가 있습니다.")
            Flag1 = True
            break

if flag1 == False:
    print("원본 파일에는 겹치는 키가 없습니다.")

flag2 = False
wb_new = openpyxl.load_workbook(r".\files\address_new.xlsx")
ws_new = wb_new.active
mc_new = ws_new.max_column
mr_new = ws_new.max_row

for i in range(2, mr_new):
    for j in range(i+1, mr_new+1):
        if ws_new.cell(i, 1).value == ws_new.cell(j, 1).value:
            print(f"비교할 파일: {i}, {j}겹치는 키가 있습니다.")
            Flag2 = True
            break

if flag2 == False:
    print("비교할 파일에는 겹치는 키가 없습니다.")
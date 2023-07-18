import openpyxl  # 패키지 import

wb = openpyxl.Workbook()  # 엑셀은 워크북(엑셀파일), 그리고 워크시트로 구분이 됩니다.

ws = wb.active  # Default Sheet를 활성화 합니다.
ws['A1'] = 'hello openpyxl'
ws['B1'] = 10
ws['B2'] = 20
ws['B3'] = '=B1+B2'

ws2 = wb.create_sheet('Hello')  # 새로운 시트를 만든다.
ws2.cell(5, 2, "My first openpyxl program")

filename = 'hello.xlsx'
wb.save(filename)

print("프로그램이 종료되었습니다.")

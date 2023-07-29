import openpyxl
from openpyxl.styles import PatternFill

wb_ori = openpyxl.load_workbook(r".\files\income_state_org.xlsx", data_only=True)
ws_ori = wb_ori.active
mc_ori = ws_ori.max_column
mr_ori = ws_ori.max_row

wb_new = openpyxl.load_workbook(r".\files\income_state_new.xlsx", data_only=True)
ws_new = wb_new.active
mc_new = ws_new.max_column
mr_new = ws_new.max_row

if mc_new != mc_ori or mr_ori != mr_new:
    print("파일의 데이터 개수에 차이가 있습니다.")

diff_color = PatternFill(start_color='ffff99', end_color='ffff99', fill_type='solid')
big_color = PatternFill(start_color='ffb399', end_color='ffb399', fill_type='solid')
small_color = PatternFill(start_color='99ffb3', end_color='99ffb3', fill_type='solid')


for i in range(1, mr_ori+1, 1):
    for j in range(1, mc_ori + 1, 1):
        if ws_ori.cell(i, j).value != ws_new.cell(i, j).value:
            ori = ws_ori.cell(i, j).value
            new = ws_new.cell(i, j).value
            ws_ori.cell(i, j).value = f"{ws_ori.cell(i, j).value} --> {ws_new.cell(i, j).value}"

            if isinstance(ori, str) or isinstance(new, str):
                ws_ori.cell(i, j).fill = diff_color
            else:
                if ori > new:
                    ws_ori.cell(i, j).fill = small_color
                else:
                    ws_ori.cell(i, j).fill = big_color

wb_ori.save(r".\files\income_state_comp.xlsx")


